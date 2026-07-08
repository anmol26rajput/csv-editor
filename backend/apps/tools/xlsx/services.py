import openpyxl
import pandas as pd
import os

class XLSXService:
    def get_workbook_structure(self, file_path):
        """Returns list of sheet names."""
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames

    def read_sheet(self, file_path, sheet_name=None, page=1, page_size=50):
        """Reads a specific sheet with pagination."""
        # Using pandas for efficient reading and JSON conversion
        # openpyxl is better for structure/editing, pandas for data reading
        
        xls = pd.ExcelFile(file_path)
        if not sheet_name:
            sheet_name = xls.sheet_names[0]
            
        # Read only necessary rows if possible (pd.read_excel doesn't support skip/limit well for random access without reading all, 
        # but we can use skiprows/nrows if we know total from metadata. For v1, read all is okay for small files)
        
        df = pd.read_excel(xls, sheet_name=sheet_name)
        
        total_rows = len(df)
        start = (page - 1) * page_size
        end = start + page_size
        
        data = df.iloc[start:end].replace({float('nan'): None}).to_dict(orient='records')
        columns = list(df.columns)
        
        return {
            'sheet_name': sheet_name,
            'columns': columns,
            'data': data,
            'total_rows': total_rows,
            'page': page,
            'page_size': page_size
        }

    def get_sheet_info(self, file_path):
        """
        Get information about all sheets in an XLSX file.
        Treats each sheet as a 'page'.
        
        Returns:
            dict: Contains total_sheets and list of sheet metadata
        """
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = []
        
        for idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            sheets.append({
                'index': idx,
                'sheet_number': idx + 1,
                'name': sheet_name,
                'rows': max_row,
                'columns': max_col
            })
        
        wb.close()
        
        return {
            'total_sheets': len(sheets),
            'sheets': sheets
        }

    def reorder_sheets(self, file_path, sheet_order, output_path):
        """
        Reorder sheets in an XLSX file.
        
        Args:
            file_path: Path to input XLSX
            sheet_order: List of sheet indices (0-indexed) in desired order
            output_path: Path for output XLSX
        
        Returns:
            output_path: Path to reordered XLSX
        """
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames
        total_sheets = len(sheet_names)
        
        # Validate sheet order
        if len(sheet_order) != total_sheets:
            raise ValueError(f"Sheet order length ({len(sheet_order)}) must match total sheets ({total_sheets})")
        
        if set(sheet_order) != set(range(total_sheets)):
            raise ValueError("Sheet order must contain all sheet indices exactly once")
        
        # Create new workbook
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # Remove default sheet
        
        # Copy sheets in new order
        for sheet_idx in sheet_order:
            source_sheet = wb[sheet_names[sheet_idx]]
            target_sheet = new_wb.create_sheet(source_sheet.title)
            
            # Copy all cells
            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet[cell.coordinate]
                    target_cell.value = cell.value
                    
                    # Copy cell formatting if needed
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()
            
            # Copy column dimensions
            for col in source_sheet.column_dimensions:
                if col in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
            
            # Copy row dimensions
            for row in source_sheet.row_dimensions:
                if row in source_sheet.row_dimensions:
                    target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
        
        new_wb.save(output_path)
        wb.close()
        new_wb.close()
        
        return output_path
